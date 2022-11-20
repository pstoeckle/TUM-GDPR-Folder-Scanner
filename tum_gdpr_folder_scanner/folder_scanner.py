"""
Logic for scanning.
"""
from itertools import chain
from json import dumps, loads
from logging import getLogger
from pathlib import Path
from re import compile as re_compile
from typing import AbstractSet, MutableMapping, MutableSet
from zipfile import BadZipFile

from click import echo
from openpyxl import load_workbook
from tqdm import tqdm

from tika import parser

_LOGGER = getLogger(__name__)
_WHITESPACE = re_compile(r"\s")
# If A is the firstname and B is the lastname, we check currently for AB, BA, A;B, B;A, etc.
# Feel free to add more separators!
_NAME_SEPARATORS: AbstractSet[str] = frozenset(["", ";", ",", "_", "-"])


class FolderScanner(object):
    """
    Scanner class.
    """

    _CACHE_FILE_NAME: str = ".scan-svn.json"
    _firstname: str
    _lastname: str
    _name_variations: AbstractSet[str]
    _matriculation_no_normalized: str
    _tum_name_normalized: str
    _files_with_name: MutableSet[Path]
    _files_that_might_contain_the_name: MutableSet[Path]
    _files_with_matriculation_no: MutableSet[Path]
    _files_with_tum_name: MutableSet[Path]
    _cache: MutableMapping[str, str] = {}

    def __init__(self, tum_id: str, name_to_search: str, matriculation_no: str) -> None:
        self._firstname, self._lastname = name_to_search.casefold().split(
            " ", maxsplit=1
        )
        self._name_variations = _create_name_variants(self._firstname, self._lastname)
        # We strip the leading zeros because some CSV do not have them for the matriculation number.
        self._matriculation_no_normalized = _normalize(matriculation_no).lstrip("0")
        self._tum_name_normalized = _normalize(tum_id)
        self._files_with_name: MutableSet[Path] = set()
        self._files_that_might_contain_the_name: MutableSet[Path] = set()
        self._files_with_matriculation_no: MutableSet[Path] = set()
        self._files_with_tum_name: MutableSet[Path] = set()
        self._cache = {}

    def scan(self, directory: Path, skip_pdfs: bool, skip_xlsx: bool) -> None:
        """
        Starts the scann process.
        :param directory:
        :param skip_pdfs:
        :param skip_xlsx:
        :return:
        """
        if not skip_pdfs:
            cache_file = directory.joinpath(self._CACHE_FILE_NAME)
            if cache_file.is_file():
                _LOGGER.info(f"Loading cache from {cache_file}")
                self._cache = loads(cache_file.read_text())
            echo("Starting the PDF scan...")
            for t in tqdm(list(directory.glob("**/*.pdf"))):
                if (normalized_text := self._cache.get(str(t))) is None:
                    _LOGGER.debug(f"{t} was not in the cache..")
                    text = parser.from_file(str(t))
                    if text["content"] is None:
                        _LOGGER.error(f"Could not extract text from {t}")
                        self._cache[str(t)] = ""
                        continue
                    normalized_text = _normalize(text["content"])
                    self._cache[str(t)] = normalized_text
                else:
                    _LOGGER.info(
                        f"We found {t} in the cache. Skip PDF2Text for this file."
                    )
                self._add_filename_to_sets(
                    normalized_text,
                    t,
                )

            echo("PDF scan: done!")

            cache_file.write_text(dumps(self._cache))
        else:
            echo("Skipping the PDFs.")
        if not skip_xlsx:
            echo("Starting the XLSX scan ...")
            for t in tqdm(list(directory.glob("**/*.xlsx"))):
                if t.stem.startswith("~"):
                    _LOGGER.warning(f"{t} DO NOT COMMIT ~$ files!")
                    continue
                try:
                    wb = load_workbook(str(t))
                except BadZipFile:
                    _LOGGER.warning(f"Could not open {t}")
                    continue

                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(sheet.min_row, sheet.max_row):
                        for i, cell in enumerate(row):
                            value = _normalize(str(cell.value))
                            self._add_filename_to_sets(
                                value,
                                t,
                            )
                            if i > 0:
                                if any(
                                    n in _normalize(str(row[i - 1].value)) + value
                                    for n in self._name_variations
                                ):
                                    self._files_with_name.add(t)
                            try:
                                next_cell = row[i + 1]
                            except IndexError:
                                continue
                            if any(
                                n in value + _normalize(str(next_cell.value))
                                for n in self._name_variations
                            ):
                                self._files_with_name.add(t)
            echo("XLSX scan: done!")
        else:
            echo("Skipping the XLSX.")
        echo("Starting the CSV and TXT scan ...")
        for t in tqdm(
            list(
                chain(
                    directory.glob("**/*.csv"),
                    directory.glob("**/*.txt"),
                    directory.glob("**/*.xml"),
                )
            )
        ):
            try:
                text = t.read_text()
            except UnicodeDecodeError:
                _LOGGER.debug(f"Encoding problem with {t}")
                encoding_found = False
                for encoding in ["iso-8859-1", "cp1252"]:
                    try:
                        text = t.read_text(encoding=encoding)
                        encoding_found = True
                        break
                    except UnicodeDecodeError:
                        _LOGGER.debug(f"Also not working with {encoding}")
                if encoding_found:
                    _LOGGER.debug(f"Encoding of {t}: {encoding}")
                else:
                    _LOGGER.warning(f"Could not find encoding for {t}")
                    continue
            except FileNotFoundError:
                _LOGGER.debug(f"File {t} could not be opened.")
                continue
            normalized_text = _normalize(text)

            self._add_filename_to_sets(
                normalized_text,
                t,
            )
        echo("CSV and TXT scan: Done!")

        _print_stuff(self._files_with_name, "name")
        _print_stuff(self._files_with_tum_name, "TUM ID")
        _print_stuff(self._files_with_matriculation_no, "matriculation number")
        if (
            len(self._files_with_name) == 0
            and len(self._files_with_tum_name) == 0
            and len(self._files_with_matriculation_no) == 0
        ):
            _print_stuff(
                self._files_that_might_contain_the_name, "firstname and lastname"
            )

    def _add_filename_to_sets(self, normalized_text: str, t: Path) -> None:
        if self._matriculation_no_normalized in normalized_text:
            self._files_with_matriculation_no.add(t)
        if self._tum_name_normalized in normalized_text:
            self._files_with_tum_name.add(t)
        if any(n in normalized_text for n in self._name_variations):
            self._files_with_name.add(t)
        elif self._firstname in normalized_text and self._lastname in normalized_text:
            self._files_that_might_contain_the_name.add(t)


def _print_stuff(files: AbstractSet[Path], label: str) -> None:
    """

    :param files:
    :param label:
    :return:
    """
    if len(files) > 0:
        echo(f"The following files contain the {label} in any order")
        for i, f in enumerate(sorted(files)):
            echo(f"{i + 1}. {f}")
    else:
        echo(f"We haven't found the {label} in any file.")


def _normalize(s: str) -> str:
    """

    :param s:
    :return:
    """
    return _WHITESPACE.sub("", s.casefold().replace('"', "").replace("'", ""))


def _create_name_variants(firstname: str, lastname: str) -> AbstractSet[str]:
    return frozenset(
        list(
            chain(
                *[
                    [firstname + sep + lastname, lastname + sep + firstname]
                    for sep in _NAME_SEPARATORS
                ]
            )
        )
        + [
            _normalize(
                f"<FAMILY_NAME_OF_STUDENT>{lastname}</FAMILY_NAME_OF_STUDENT><FIRST_NAME_OF_STUDENT>{firstname}</FIRST_NAME_OF_STUDENT>"
            ),
            _normalize(
                f"<FIRST_NAME_OF_STUDENT>{firstname}</FIRST_NAME_OF_STUDENT><FAMILY_NAME_OF_STUDENT>{lastname}</FAMILY_NAME_OF_STUDENT>"
            ),
        ]
    )
