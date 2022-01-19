"""
Script to find students.
"""
from itertools import chain
from logging import INFO, basicConfig, getLogger
from pathlib import Path
from re import compile as re_compile
from sys import stdout
from typing import AbstractSet, MutableSet, Optional
from zipfile import BadZipFile

from openpyxl import load_workbook
from tqdm import tqdm

from scan_svn import __version__
from tika import parser
from typer import Exit, Option, Typer, echo

basicConfig(
    format="%(levelname)s: %(asctime)s: %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    level=INFO,
    stream=stdout,
)
_LOGGER = getLogger(__name__)
app = Typer()

_WHITESPACE = re_compile(r"\s")


def _version_callback(value: bool) -> None:
    """

    :param value:
    :return:
    """
    if value:
        echo(f"scan-svn {__version__}")
        raise Exit()


@app.command()
def scan_directory(
    directory: Path = Option(
        ".",
        "--directory",
        "-s",
        help="The directory we want to analyze.",
        file_okay=False,
        resolve_path=True,
    ),
    name_to_search: str = Option(
        None,
        "--name-to-search",
        "-n",
        prompt=True,
        help="The name we are looking for. Please write the name in the form `Lastname Firstname` or `Firstname Lastname`.",
    ),
    matriculation_no: str = Option(
        None,
        "--matriculation-no",
        "-m",
        prompt=True,
        help="The matriculation number we are looking for.",
    ),
    tum_id: str = Option(
        None, "--tum-id", "-t", prompt=True, help="The TUM ID, e.g., ga12acb."
    ),
    skip_pdfs: bool = Option(
        False,
        "--skip-pdfs",
        "-S",
        is_flag=True,
        help="The PDF extraction takes some time. You can skip it for a first run.",
    ),
    skip_xlsx: bool = Option(
        False,
        "--skip-xlsx",
        "-X",
        is_flag=True,
        help="The XLSX extraction takes some time. You can skip it for a first run.",
    ),
    _: Optional[bool] = Option(
        None,
        "--version",
        "-v",
        callback=_version_callback,
        help="Shows the version and exits.",
    ),
) -> None:
    """
    Scans all relevant files (CSV, PDF, TXT, XLSX, XML) for the given name, TUM name, and matriculation number.
    """
    name_variations = _create_name_variants(name_to_search)
    # We strip the leading zeros because some CSV do not have them for the matriculation number.
    matriculation_no_normalized = _normalize(matriculation_no).lstrip("0")
    tum_name_normalized = _normalize(tum_id)

    files_with_name: MutableSet[Path] = set()
    files_with_matriculation_no: MutableSet[Path] = set()
    files_with_tum_name: MutableSet[Path] = set()
    if not skip_pdfs:
        echo("Starting the PDF scan...")
        for t in tqdm(list(directory.glob("**/*.pdf"))):
            text = parser.from_file(str(t))
            if text["content"] is None:
                _LOGGER.debug(f"Could not extract text from {t}")
                continue
            normalized_text = _normalize(text["content"])
            _add_filename_to_sets(
                files_with_matriculation_no,
                files_with_name,
                files_with_tum_name,
                matriculation_no_normalized,
                name_variations,
                normalized_text,
                t,
                tum_name_normalized,
            )
        echo("PDF scan: done!")
    else:
        echo("Skipping the PDFs.")
    if not skip_xlsx:
        echo("Starting the XLSX scan ...")
        for t in tqdm(list(directory.glob("**/*.xlsx"))):
            if t.stem.startswith("~"):
                _LOGGER.debug(f"{t} DO NOT COMMIT ~$ files!")
                continue
            try:
                wb = load_workbook(str(t))
            except BadZipFile:
                _LOGGER.warning(f"Could not open {t}")
                continue

            for sheet in wb.worksheets:
                for row in sheet.iter_rows(sheet.min_row, sheet.max_row):
                    for i, cell in enumerate(row):
                        value = _normalize(str(cell.value))
                        _add_filename_to_sets(
                            files_with_matriculation_no,
                            files_with_name,
                            files_with_tum_name,
                            matriculation_no_normalized,
                            name_variations,
                            value,
                            t,
                            tum_name_normalized,
                        )
                        if i > 0:
                            if any(
                                n in _normalize(str(row[i - 1].value)) + value
                                for n in name_variations
                            ):
                                files_with_name.add(t)
                        try:
                            next_cell = row[i + 1]
                        except IndexError:
                            continue
                        if any(
                            n in value + _normalize(str(next_cell.value))
                            for n in name_variations
                        ):
                            files_with_name.add(t)
        echo("XLSX scan: done!")
    else:
        echo("Skipping the XLSX.")
    echo("Starting the CSV and TXT scan ...")
    for t in tqdm(
        list(
            chain(
                directory.glob("**/*.csv"),
                directory.glob("**/*.txt"),
                directory.glob("**/*.xml"),
            )
        )
    ):
        try:
            text = t.read_text()
        except UnicodeDecodeError:
            _LOGGER.debug(f"Encoding problem with {t}")
            encoding_found = False
            for encoding in ["iso-8859-1", "cp1252"]:
                try:
                    text = t.read_text(encoding=encoding)
                    encoding_found = True
                    break
                except UnicodeDecodeError:
                    _LOGGER.debug(f"Also not working with {encoding}")
            if encoding_found:
                _LOGGER.debug(f"Encoding of {t}: {encoding}")
            else:
                _LOGGER.warning(f"Could not find encoding for {t}")
                continue
        except FileNotFoundError:
            _LOGGER.debug(f"File {t} could not be opened.")
            continue
        normalized_text = _normalize(text)

        _add_filename_to_sets(
            files_with_matriculation_no,
            files_with_name,
            files_with_tum_name,
            matriculation_no_normalized,
            name_variations,
            normalized_text,
            t,
            tum_name_normalized,
        )
    echo("CSV and TXT scan: Done!")

    _print_stuff(files_with_name, "name")
    _print_stuff(files_with_tum_name, "TUM ID")
    _print_stuff(files_with_matriculation_no, "matriculation number")


def _add_filename_to_sets(
    files_with_matriculation_no,
    files_with_name,
    files_with_tum_name,
    matriculation_no_normalized,
    name_variations,
    normalized_text,
    t,
    tum_name_normalized,
):
    if matriculation_no_normalized in normalized_text:
        files_with_matriculation_no.add(t)
    if tum_name_normalized in normalized_text:
        files_with_tum_name.add(t)
    if any(n in normalized_text for n in name_variations):
        files_with_name.add(t)


def _print_stuff(files: AbstractSet[Path], label: str) -> None:
    """

    :param files:
    :param label:
    :return:
    """
    if len(files) > 0:
        echo(f"The following files contain the {label} in any order")
        for i, f in enumerate(sorted(files)):
            echo(f"{i + 1}. {f}")
    else:
        echo(f"We haven't found the {label} in any file.")


def _normalize(s: str) -> str:
    """

    :param s:
    :return:
    """
    return _WHITESPACE.sub("", s.casefold().replace('"', "").replace("'", ""))


def _create_name_variants(name: str) -> AbstractSet[str]:
    firstname, lastname = name.casefold().split(" ", maxsplit=1)
    return frozenset(
        [
            firstname + lastname,
            lastname + firstname,
            firstname + "," + lastname,
            lastname + "," + firstname,
            firstname + ";" + lastname,
            lastname + ";" + firstname,
            _normalize(
                f"<FAMILY_NAME_OF_STUDENT>{lastname}</FAMILY_NAME_OF_STUDENT><FIRST_NAME_OF_STUDENT>{firstname}</FIRST_NAME_OF_STUDENT>"
            ),
            _normalize(
                f"<FIRST_NAME_OF_STUDENT>{firstname}</FIRST_NAME_OF_STUDENT><FAMILY_NAME_OF_STUDENT>{lastname}</FAMILY_NAME_OF_STUDENT>"
            ),
        ]
    )


if __name__ == "__main__":
    app()
